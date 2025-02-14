#! python3
import sys

import openpyxl

def cellInverter(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    newSheet = wb.create_sheet('new')
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            newSheet.cell(column=i, row=j).value = sheet.cell(column=j, row=i).value
    wb.save(file)

if len(sys.argv) == 2:
    cellInverter(sys.argv[1])