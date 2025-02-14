#! python3
# multiplicationTable.py - 创建 N*N 的乘法表格

import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string


def main(n):
    wb = openpyxl.Workbook()
    sheet = wb.active
    boldFont = Font(bold=True)
    for i in range(1, n + 2):
        for j in range(1, n + 2):
            # col = get_column_letter(j)
            # coord = col + str(i)
            cell = sheet.cell(row=i, column=j)

            if i == 1 or j == 1:
                cell.font = boldFont

            if not (i == 1 and j == 1):
                cell.value = max(1, i - 1) * max(1, j - 1)

    wb.save('table%s.xlsx' % n)
    print('table%s.xlsx' % n)

if len(sys.argv) == 2:
    main(int(sys.argv[1]))

