#! python3
# unmergeCells.py - Unmerges cells in a spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3')  # Unmerge all these cells.
sheet.unmerge_cells('C5:D5')  # Unmerge these two cells.
wb.save('merged.xlsx')
